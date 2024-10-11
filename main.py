import eel
import pandas as pd
from openpyxl import load_workbook
import re
import os
import base64
from io import BytesIO
import datetime
import zipfile
import logging

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 初始化全局变量
edi_content = ""
results_dict = {}
filtered_results = {}
edi_save_directory = ""
marked_boxes = {"LB": [], "AC": []}
success_count = 0
cleaned_count = 0
addition_count = 0
excel_io = None

# 初始化Eel
eel.init('web')

def clean_symbols(text):
    global cleaned_count
    if text is None:
        return None
    text_str = str(text).replace(" ", "")

    # 去除括号及括号内的内容
    original_text = text_str
    text_str = re.sub(r'\（.*?\）|\(.*?\)', '', text_str)
    if text_str != original_text:
        cleaned_count += 1  # 记录一次清理操作

    if '.' in text_str:
        # 保留数字和小数点
        numeric_str = re.sub(r'[^\d.]', '', text_str)
        try:
            return float(numeric_str) if numeric_str else None
        except ValueError:
            return None
    else:
        # 只保留数字
        numeric_str = re.sub(r'\D', '', text_str)
        try:
            return int(numeric_str) if numeric_str else None
        except ValueError:
            return None

def clean_box_number(text):
    if text is None:
        return None
    # 只移除前后的空白字符，保留箱号中的所有字符
    return str(text).strip()

@eel.expose
def import_excel(file_content_base64):
    global excel_io, results_dict, cleaned_count, filtered_results
    cleaned_count = 0
    if not file_content_base64:
        eel.display_message("未选择Excel文件。")
        return

    try:
        import openpyxl
        file_bytes = base64.b64decode(file_content_base64)
        excel_io = BytesIO(file_bytes)

        # 尝试打开文件
        try:
            wb = load_workbook(filename=excel_io, data_only=True)
        except zipfile.BadZipFile:
            eel.display_message("错误：所选的Excel文件可能版本过旧。请确保您选择了.xlsx格式的文件。")
            return
        except Exception as e:
            eel.display_message(f"打开Excel文件时出错：{str(e)}")
            return

        if 'DG' not in [sheet.upper() for sheet in wb.sheetnames]:
            eel.display_message("Excel文件中没有名为'DG'的工作表。")
            return
        ws = wb[next(sheet for sheet in wb.sheetnames if sheet.upper() == 'DG')]

        # 检测列索引
        header_row = None
        cntr_col = None
        imdg_col = None
        un_col = None

        for row in ws.iter_rows(min_row=1, max_row=10, max_col=ws.max_column):
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value).upper()  # 转换为大写
                    if re.search(r'CNTR\s*NO\.?', cell_value):
                        cntr_col = cell.column
                        header_row = cell.row
                    elif 'IMDG' in cell_value:
                        imdg_col = cell.column
                    elif re.search(r'UN\s*NO\.?', cell_value):
                        un_col = cell.column

            if header_row and cntr_col and imdg_col and un_col:
                break

        if not (header_row and cntr_col and imdg_col and un_col):
            eel.display_message("无法在Excel中找到必要的列。")
            return
        
        eel.display_message(f"找到列：CNTR NO.={cntr_col}, IMDG={imdg_col}, UN NO={un_col}, 表头行={header_row}")

        results_dict = {}
        merged_ranges = ws.merged_cells.ranges

        for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
            box_number = None  
            for merged_range in merged_ranges:
                min_col = getattr(merged_range, 'min_col', None) or merged_range.bounds[0]
                min_row = getattr(merged_range, 'min_row', None) or merged_range.bounds[1]
                max_row = getattr(merged_range, 'max_row', None) or merged_range.bounds[3]
                
                if min_col == cntr_col and min_row <= row[0].row <= max_row:
                    box_number = clean_box_number(ws.cell(row=min_row, column=cntr_col).value)
                    break
            
            if not box_number:
                box_number = clean_box_number(row[cntr_col-1].value)

            if not box_number:
                continue

            i_value = clean_symbols(row[imdg_col-1].value)
            j_value = clean_symbols(row[un_col-1].value)

            if i_value is not None and j_value is not None:
                if box_number not in results_dict:
                    results_dict[box_number] = []
                results_dict[box_number].append({
                    'I': i_value,
                    'J': j_value
                })

        filtered_results = {k: v for k, v in results_dict.items() if len(v) > 1}

        if not filtered_results:
            eel.display_message("没有找到符合条件的箱号。")
            return

        eel.display_message("Excel文件导入成功。")
        eel.display_message(f"', '.join(filtered_results.keys())")
        eel.display_message(f"已修正的字符串个数：{cleaned_count}")

        eel.show_filtered_results(filtered_results)
        eel.display_message("单箱装运多类危险品的箱号信息获取成功！")

        # 添加对 Special 工作表的处理
        process_special_containers()

    except Exception as e:
        logging.error(f"导入Excel文件时出错：{str(e)}")
        eel.display_message(f"导入Excel文件时出错：{str(e)}")

@eel.expose
def import_edi(file_content, save_directory):
    global edi_content, edi_save_directory
    if not file_content:
        eel.display_message("未选择EDI报文文件。")
        return
    try:
        edi_content = file_content
        edi_save_directory = save_directory  # 保存目录

        # 检查中文字符
        if any('\u4e00' <= char <= '\u9fa5' for char in edi_content):
            # 找到所有中文字符并列出
            illegal_chars = set(char for char in edi_content if '\u4e00' <= char <= '\u9fa5')
            illegal_chars_str = ', '.join(illegal_chars)
            eel.display_message(f"报文中含有非法字符：{illegal_chars_str}")
            return

        eel.display_message("EDI报文文件已成功导入。")
    except Exception as e:
        eel.display_message(f"导入EDI报文文件时出错：{e}")

@eel.expose
def execute_modify_and_mark():
    global edi_content, filtered_results
    try:
        if not edi_content:
            eel.display_message("请先导入EDI报文文件。")
            return
        if not filtered_results:
            eel.display_message("请先导入并处理Excel文件。")
            return

        eel.display_message("开始执行修改和标记操作...")
        modify_edi()
        lithium_containers, carbon_containers = process_special_containers()
        
        eel.clear_special_log()
        
        lithium_success_count = 0
        eel.js_display_special_message("锂电池箱号标记结果:")
        for index, container in enumerate(lithium_containers, 1):
            status = mark_special_container(container, "LB")
            eel.update_special_log(f"{index}. {container}", status)
            if status == "标记完成":
                lithium_success_count += 1
        
        eel.js_display_special_message(f"\n锂电池箱号成功标记次数: {lithium_success_count}")
        
        carbon_success_count = 0
        eel.js_display_special_message("\n碳产品箱号标记结果:")
        for index, container in enumerate(carbon_containers, 1):
            status = mark_special_container(container, "AC")
            eel.update_special_log(f"{index}. {container}", status)
            if status == "标记完成":
                carbon_success_count += 1
        
        eel.js_display_special_message(f"\n碳产品箱号成功标记次数: {carbon_success_count}")
        
        save_edi()
        eel.display_message("EDI报文处理完成，已保存修改后的文件。")
    except Exception as e:
        logging.error(f"执行修改和标记过程中出错：{str(e)}")
        eel.display_message(f"执行修改和标记过程中出错：{str(e)}")
    finally:
        eel.enable_modify_and_mark_button()

def modify_edi():
    global edi_content, filtered_results, addition_count
    addition_count = 0

    if not edi_content or not filtered_results:
        return

    eel.display_message("开始处理拼箱危险品箱号：")
    for index, (box_number, items) in enumerate(filtered_results.items(), 1):
        box_pattern = re.escape(box_number)
        matches = list(re.finditer(box_pattern, edi_content))

        if matches:
            eel.update_main_log(f"{index}. {box_number}: 处理中")
            for match in matches:
                dgs_imd_pattern = r'DGS\+IMD\+'
                search_start = match.end()
                dgs_imd_match = re.search(dgs_imd_pattern, edi_content[search_start:])

                if dgs_imd_match:
                    dgs_imd_end = search_start + dgs_imd_match.end()
                    apostrophe_match = re.search(r"'", edi_content[dgs_imd_end:])

                    if apostrophe_match:
                        apostrophe_end = dgs_imd_end + apostrophe_match.end()
                        insert_position = apostrophe_end

                        existing_content = edi_content[dgs_imd_end:apostrophe_end - 1]
                        existing_parts = existing_content.split('+')
                        existing_A = None
                        existing_B = None
                        if len(existing_parts) >= 2:
                            try:
                                existing_A = float(existing_parts[-2]) if '.' in existing_parts[-2] else int(existing_parts[-2])
                                existing_B = float(existing_parts[-1]) if '.' in existing_parts[-1] else int(existing_parts[-1])
                            except ValueError:
                                pass

                        for item in items:
                            imdg = item['I']
                            un = item['J']

                            if existing_A == imdg and existing_B == un:
                                continue

                            insert_str = f"DGS+IMD+{imdg}+{un}'"

                            existing_str = edi_content[insert_position:insert_position + len(insert_str)]
                            if existing_str == insert_str:
                                continue
                            else:
                                edi_content = edi_content[:insert_position] + insert_str + edi_content[insert_position:]
                                insert_position += len(insert_str)
                                addition_count += 1

            status = f"副危险性信息添加成功 (添加 {len(items)} 条)"
        else:
            status = "未在EDI报文中找到"

        eel.update_main_log(f"{index}. {box_number}: {status}")

    eel.display_message(f"拼箱危险品箱号处理完成，实际添加的副危险性信息次数：{addition_count}")

@eel.expose
def mark_as_lb(file_content):
    mark_container("LB", file_content)

@eel.expose
def mark_as_ac(file_content):
    mark_container("AC", file_content)

def mark_container(label, file_content):
    global marked_boxes
    if not file_content:
        eel.display_message(f"未选择标记为{label}的箱号文件。")
        return
    try:
        container_numbers = file_content.splitlines()

        for container in container_numbers:
            if container.strip():  # 忽略空行
                marked_boxes[label].append(container.strip())

        eel.display_message(f"箱号成功导入为{label}标记！")
    except Exception as e:
        eel.display_message(f"导入标记为{label}的箱号文件时出错：{e}")

def mark_containers(label):
    global edi_content, success_count
    containers = marked_boxes[label]
    new_edi_data = edi_content.splitlines()

    for i, line in enumerate(new_edi_data):
        for container in containers:
            if container in line:
                container_pos = line.find(container) + len(container)
                remaining_line = line[container_pos:]
                first_quote_index = remaining_line.find("'")

                if first_quote_index != -1:
                    insert_str = f"FTX+AAA+++{label}'FTX+HAN+++OD'"
                    insert_position = container_pos + first_quote_index + 1

                    existing_str = line[insert_position: insert_position + len(insert_str)]
                    if existing_str == insert_str:
                        continue
                    else:
                        line = line[:insert_position] + insert_str + line[insert_position:]
                        new_edi_data[i] = line
                        success_count += 1

    edi_content = "\n".join(new_edi_data)

@eel.expose
def execute_marking():
    global success_count
    success_count = 0

    mark_containers("LB")
    mark_containers("AC")

    eel.display_message(f"成功标记特殊积载的箱号数量：{success_count}")

@eel.expose
def save_edi():
    global edi_content, edi_save_directory
    if not edi_content:
        eel.display_message("没有EDI内容可保存。")
        return

    try:
        if not edi_save_directory:
            eel.display_message("未指定保存EDI报文文件的目录。")
            return

        os.makedirs(edi_save_directory, exist_ok=True)

        current_date = datetime.datetime.now().strftime("%m%d%H%M")
        new_file_name = f"modifiedEDI_{current_date}.edi"
        new_file_path = os.path.join(edi_save_directory, new_file_name)

        with open(new_file_path, 'w', encoding='utf-8') as file:
            file.write(edi_content)
        eel.display_message(f"修改后的EDI报文已保存到 {new_file_path}")
    except Exception as e:
        eel.display_message(f"保存EDI报文文件时出错：{e}")

@eel.expose
def reset():
    global edi_content, results_dict, filtered_results, edi_save_directory, marked_boxes, success_count, cleaned_count, addition_count, excel_io
    edi_content = ""
    results_dict = {}
    filtered_results = {}
    edi_save_directory = ""
    marked_boxes = {"LB": [], "AC": []}
    success_count = 0
    cleaned_count = 0
    addition_count = 0
    excel_io = None  # 重置 excel_io
    logging.info("程序状态已重置")
    eel.display_message("所有数据已重置。")
    eel.reset_display()
    eel.display_message("程序已重置，准备就绪。")

@eel.expose
def get_edi_content():
    return edi_content

def process_special_sheet(ws):
    lithium_battery_containers = []
    carbon_products_containers = []
    
    lithium_section = False
    carbon_section = False
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        row_content = str(row[0]).strip().upper() if row[0] else ""
        
        if "LITHIUM BATTERY SHIPMENT LIST" in row_content:
            lithium_section = True
            carbon_section = False
            continue
        elif "CARBON PRODUCTS LIST" in row_content:
            lithium_section = False
            carbon_section = True
            continue
        elif "EXCEPT FOR" in row_content:
            break  # 直接结束处理，忽略后续内容
        
        cntr_no = clean_and_validate_container(row[6])  # G列
        if cntr_no:
            if lithium_section:
                lithium_battery_containers.append(cntr_no)
            elif carbon_section:
                carbon_products_containers.append(cntr_no)
    
    return lithium_battery_containers, carbon_products_containers

def clean_and_validate_container(cntr_no):
    if not cntr_no:
        return None
    
    # 清理箱号，只保留字母和数字
    cleaned = re.sub(r'[^A-Za-z0-9]', '', str(cntr_no))
    
    # 验证箱号格式
    if re.match(r'^[A-Za-z]{4}\d{7}$', cleaned):
        return cleaned
    return None  # 如果不符合标准格式，返回None

@eel.expose
def process_special_containers():
    global edi_content
    
    try:
        wb = load_workbook(filename=excel_io, data_only=True)
        if 'SPECIAL' not in [sheet.upper() for sheet in wb.sheetnames]:
            eel.js_display_special_message("Excel文件中没有名为'Special'的工作表。")
            return [], []
        
        ws = wb[next(sheet for sheet in wb.sheetnames if sheet.upper() == 'SPECIAL')]
        lithium_containers, carbon_containers = process_special_sheet(ws)
        
        eel.js_display_special_message("需要标记的锂电池箱号:")
        for index, container in enumerate(lithium_containers, 1):
            eel.update_special_log(f"{index}. {container}", "待标记")
        
        eel.js_display_special_message(f"\n共 {len(lithium_containers)} 个锂电池箱号")
        
        eel.js_display_special_message("\n需要标记的碳产品箱号:")
        for index, container in enumerate(carbon_containers, 1):
            eel.update_special_log(f"{index}. {container}", "待标记")
        
        eel.js_display_special_message(f"\n共 {len(carbon_containers)} 个碳产品箱号")
        
        eel.js_display_special_message("\n注意：'其他特殊积载'的内容已被忽略。")
        
        return lithium_containers, carbon_containers
        
    except Exception as e:
        eel.js_display_special_message(f"处理特殊积载箱号时出错: {str(e)}")
        logging.error(f"处理特殊积载箱号时出错: {str(e)}")
        return [], []

def mark_special_container(container, label):
    global edi_content
    
    if container not in edi_content:
        return "未在EDI报文中找到"
    
    pattern = re.escape(container) + r".*?'"
    match = re.search(pattern, edi_content)
    if not match:
        return "在EDI报文中找到但无法标记"
    
    insert_position = match.end()
    insert_str = f"FTX+AAA+++{label}'FTX+HAN+++OD'"
      
    
    edi_content = edi_content[:insert_position] + insert_str + edi_content[insert_position:]
    return "标记完成"

@eel.expose
def display_special_message(message):
    eel.js_display_special_message(message)  # 调用 JavaScript 函数 
    print(f"Special: {message}")  # 用于调试

@eel.expose
def update_special_log(container, status):
    eel.js_display_special_message(f"{container}: {status}")

@eel.expose
def show_filtered_results(filtered_results):
    eel.display_message("拼箱危险品的箱号清单：")
    for index, (box, items) in enumerate(filtered_results.items(), 1):
        eel.update_main_log(f"{index}. {box}: 待处理")
        for item_index, item in enumerate(items, 1):
            eel.display_message(f"  {index}.{item_index} IMDG: {item['I']}, UN: {item['J']}")
    eel.display_message(f"共发现 {len(filtered_results)} 个拼箱危险品箱号")

# 启动Eel应用
if __name__ == '__main__':
    try:
        eel.start('index.html', size=(1000, 800), port=8000, mode='chrome')
    except EnvironmentError:
        # 如果 Chrome 不可用或端口 8000 被占用，尝试使用其他端口
        import socket
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.bind(('', 0))
        port = sock.getsockname()[1]
        sock.close()
        eel.start('index.html', size=(1000, 800), port=port, mode='default')